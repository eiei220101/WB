from __future__ import annotations

import json
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.drawing.image import Image as XLImage


@dataclass(frozen=True)
class ExportArtifacts:
    filled_xlsx_path: Path
    pdf_path: Path | None


def _iter_registry_install_locations() -> list[Path]:
    """
    LibreOffice のインストール場所をレジストリから探す（Windows向け）。
    """
    try:
        import winreg  # type: ignore
    except Exception:
        return []

    paths: list[Path] = []

    def _walk_uninstall(root, view_flag) -> None:
        try:
            base = winreg.OpenKey(root, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall", 0, winreg.KEY_READ | view_flag)
        except OSError:
            return
        try:
            i = 0
            while True:
                try:
                    sub_name = winreg.EnumKey(base, i)
                except OSError:
                    break
                i += 1
                try:
                    sub = winreg.OpenKey(base, sub_name, 0, winreg.KEY_READ | view_flag)
                except OSError:
                    continue
                try:
                    disp, _ = winreg.QueryValueEx(sub, "DisplayName")
                    if not isinstance(disp, str) or "LibreOffice" not in disp:
                        continue
                    loc, _ = winreg.QueryValueEx(sub, "InstallLocation")
                    if isinstance(loc, str) and loc.strip():
                        paths.append(Path(loc.strip()))
                except OSError:
                    continue
        finally:
            try:
                winreg.CloseKey(base)
            except OSError:
                pass

    # 64bit / 32bit view
    HKLM = getattr(__import__("winreg"), "HKEY_LOCAL_MACHINE")
    HKCU = getattr(__import__("winreg"), "HKEY_CURRENT_USER")
    KEY_WOW64_64KEY = getattr(__import__("winreg"), "KEY_WOW64_64KEY", 0)
    KEY_WOW64_32KEY = getattr(__import__("winreg"), "KEY_WOW64_32KEY", 0)
    for root in (HKLM, HKCU):
        _walk_uninstall(root, KEY_WOW64_64KEY)
        _walk_uninstall(root, KEY_WOW64_32KEY)

    # de-dupe while keeping order
    seen: set[str] = set()
    out: list[Path] = []
    for p in paths:
        s = str(p).lower()
        if s in seen:
            continue
        seen.add(s)
        out.append(p)
    return out


def _repo_templates_dir() -> Path:
    return Path(__file__).resolve().parent / "templates"


def template_path_for_tail(tail: str) -> Path:
    # Convention: templates/<TAIL>_template.xlsx
    return _repo_templates_dir() / f"{tail}_template.xlsx"


def mapping_path_for_tail(tail: str) -> Path:
    # Convention: templates/<TAIL>_mapping.json
    return _repo_templates_dir() / f"{tail}_mapping.json"


def load_mapping(tail: str) -> dict[str, Any]:
    p = mapping_path_for_tail(tail)
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


def _which_soffice() -> str | None:
    # 0) explicit override
    override = os.environ.get("SOFFICE_PATH")
    if override:
        p = Path(override)
        if p.exists():
            return str(p)

    # 1) PATH
    for name in ("soffice", "soffice.exe", "soffice.com"):
        exe = shutil.which(name)
        if exe:
            return exe

    # 2) Common Windows install locations
    candidates: list[Path] = []
    for base in [
        os.environ.get("PROGRAMFILES"),
        os.environ.get("PROGRAMFILES(X86)"),
    ]:
        if not base:
            continue
        candidates.extend(
            [
                Path(base) / "LibreOffice" / "program" / "soffice.exe",
                Path(base) / "LibreOffice" / "program" / "soffice.com",
            ]
        )
    # 3) Hard-coded fallbacks (in case env vars are missing)
    candidates.extend(
        [
            Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
            Path(r"C:\Program Files\LibreOffice\program\soffice.com"),
            Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
            Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.com"),
        ]
    )

    # 4) Registry (InstallLocation)
    for base in _iter_registry_install_locations():
        candidates.extend(
            [
                base / "program" / "soffice.exe",
                base / "program" / "soffice.com",
                base / "soffice.exe",
                base / "soffice.com",
            ]
        )
    for c in candidates:
        if c.exists():
            return str(c)
    return None


def convert_xlsx_to_pdf(xlsx_path: Path, out_dir: Path) -> Path:
    soffice = _which_soffice()
    if not soffice:
        pf = os.environ.get("PROGRAMFILES")
        pfx86 = os.environ.get("PROGRAMFILES(X86)")
        sp = os.environ.get("SOFFICE_PATH")
        raise RuntimeError(
            "LibreOffice（soffice）が見つかりません。LibreOffice をインストールするか、PATH に soffice を通してください。\n"
            f"PROGRAMFILES={pf!r}\nPROGRAMFILES(X86)={pfx86!r}\nSOFFICE_PATH={sp!r}"
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--norestore",
        "--convert-to",
        "pdf",
        "--outdir",
        str(out_dir),
        str(xlsx_path),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(
            "LibreOffice によるPDF変換に失敗しました。\n"
            f"stdout:\n{proc.stdout}\n\nstderr:\n{proc.stderr}"
        )

    pdf = out_dir / (xlsx_path.stem + ".pdf")
    if not pdf.exists():
        raise RuntimeError("PDFが生成されませんでした（LibreOfficeの出力を確認してください）。")
    return pdf


def fill_template_xlsx(
    *,
    tail: str,
    values: dict[str, Any],
    images: dict[str, bytes] | None = None,
    work_dir: Path,
) -> ExportArtifacts:
    """
    values: 「キー -> 値」を渡す（例: {"zfm_weight": 1650.0, ...}）
    mapping: templates/<TAIL>_mapping.json で「キー -> {sheet, cell, number_format?}」を定義
    """
    tpl = template_path_for_tail(tail)
    if not tpl.exists():
        raise FileNotFoundError(f"テンプレが見つかりません: {tpl}")

    mapping = load_mapping(tail)
    if not mapping:
        raise FileNotFoundError(
            f"セル対応表が見つかりません: {mapping_path_for_tail(tail)}（まずは対応表JSONを作成してください）"
        )

    wb = openpyxl.load_workbook(tpl)

    images = images or {}

    for key, spec in mapping.items():
        if not isinstance(spec, dict):
            continue
        sheet = spec.get("sheet")
        cell = spec.get("cell")
        if not sheet or not cell:
            continue
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]

        if spec.get("type") == "image":
            if key not in images:
                continue
            img_bytes = images[key]
            img_path = work_dir / f"{tail}_{key}.png"
            img_path.write_bytes(img_bytes)
            xl_img = XLImage(str(img_path))
            wpx = spec.get("width_px")
            hpx = spec.get("height_px")
            if isinstance(wpx, (int, float)) and wpx > 0:
                xl_img.width = int(wpx)
            if isinstance(hpx, (int, float)) and hpx > 0:
                xl_img.height = int(hpx)
            ws.add_image(xl_img, cell)
        else:
            if key not in values:
                continue
            v = values[key]
            # ブラウザ表示のまま反映したいので、文字列はテキストとして書き込む
            if isinstance(v, str):
                try:
                    ws[cell].number_format = "@"
                except Exception:
                    pass
                ws[cell].value = v
            else:
                ws[cell].value = v
            nf = spec.get("number_format")
            if nf:
                ws[cell].number_format = str(nf)

    work_dir.mkdir(parents=True, exist_ok=True)
    filled = work_dir / f"{tail}_filled.xlsx"
    wb.save(filled)
    return ExportArtifacts(filled_xlsx_path=filled, pdf_path=None)


def build_print_pdf_from_template(
    *,
    tail: str,
    values: dict[str, Any],
    images: dict[str, bytes] | None = None,
) -> tuple[bytes, str]:
    """
    Returns: (pdf_bytes, filename)
    """
    with tempfile.TemporaryDirectory(prefix="wb_export_") as td:
        work_dir = Path(td)
        artifacts = fill_template_xlsx(tail=tail, values=values, images=images, work_dir=work_dir)
        pdf_path = convert_xlsx_to_pdf(artifacts.filled_xlsx_path, out_dir=work_dir)
        pdf_bytes = pdf_path.read_bytes()
        return pdf_bytes, f"{tail}_WB.pdf"

